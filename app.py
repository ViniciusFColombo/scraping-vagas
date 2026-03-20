from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

navegador = webdriver.Chrome()

endereco_principal = "https://br.indeed.com/jobs?q="
nome_vaga = "Python"
local = "&l=Remoto"

pesquisar = endereco_principal + nome_vaga + local

print(pesquisar)

navegador.get(pesquisar)

WebDriverWait(navegador, 10).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, ".job_seen_beacon"))
)

vagas = navegador.find_elements(By.CSS_SELECTOR, ".job_seen_beacon")

print("Quantidade de vagas encontradas:", len(vagas))

vagas_encontradas = []

data_coleta = datetime.now().strftime("%d/%m/%Y")

for vaga in vagas:

    try:
        titulo = vaga.find_element(By.CSS_SELECTOR, "h2").text
        empresa = vaga.find_element(By.CSS_SELECTOR, "[data-testid='company-name']").text
        local = vaga.find_element(By.CSS_SELECTOR, "[data-testid='text-location']").text
        id_vaga = vaga.find_element(By.CSS_SELECTOR, "h2 a").get_attribute("data-jk")

        link = f"https://br.indeed.com/viewjob?jk={id_vaga}"

        if titulo and empresa:
            vagas_encontradas.append({
                "Título da Vaga": titulo.strip(),
                "Empresa": empresa.strip(),
                "Local": local.strip(),
                "Data da Coleta": data_coleta,
                "Link": link
            })

    except Exception as e:
        print(f"Erro ao processar vaga: {e}")

df = pd.DataFrame(vagas_encontradas)

df.drop_duplicates(subset=["Título da Vaga", "Empresa"], inplace=True)

df.sort_values(by="Empresa", inplace=True)

df.reset_index(drop=True, inplace=True)

df.to_excel("VagasIndeed.xlsx", index=False)

arquivo_excel = "VagasIndeed.xlsx"

wb = load_workbook(arquivo_excel)
ws = wb.active

cor_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
fonte_branca = Font(color="FFFFFF", bold=True)

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = cor_cabecalho
    cell.font = fonte_branca
    cell.alignment = Alignment(horizontal="center")

for col in ws.columns:
    max_lenght = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        try:
            if cell.value:
                max_lenght = max(max_lenght, len(str(cell.value)))
        except:
            pass
    
    ws.column_dimensions[col_letter].width = max_lenght + 5

    ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"

    for row in range(2, ws.max_row +1):
        cell = ws.cell(row=row, column=5)
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

wb.save(arquivo_excel)

print("Arquivo Excel gerado com sucesso!")

navegador.quit()